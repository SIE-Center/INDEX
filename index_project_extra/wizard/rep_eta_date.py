# -*- coding: utf-8 -*-
import logging
import base64
from odoo import models, fields, api
from odoo.exceptions import ValidationError, UserError, RedirectWarning
from tempfile import NamedTemporaryFile
from datetime import date, timedelta
from odoo.tools.translate import _
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.chart import   Reference,BarChart
from io import BytesIO 
import base64
_logger = logging.getLogger(__name__)

class Index_Eta_date(models.TransientModel):
    _name = 'index_project_extra.rep_eta_date'
    _description = "Reporte de Tareas"
    sdate = fields.Datetime('Desde')
    edate = fields.Datetime('Hasta')
    reporte = fields.Selection([
        ('1', 'Facturación'),
        ('2', 'Presidencia'),
        ('3', 'Estadísticas Mensuales'),
        ('4', 'General'),
        ('5', 'Incumplimiento')
    ], string='Reporte')

    def generate(self):
        wb = Workbook() #creamos objeto
        filename = ''
        cand = False
        if self.reporte == '1':#IMMEX
            cand = self.env['custom.vlines'].search([('eta_date','>=',self.sdate),('eta_date','<=',self.edate)])
        if self.reporte == '2':#Presidencia
            cand = self.env['custom.vlines'].search([('eta_date','>=',self.sdate),('eta_date','<=',self.edate),('etapa','in',('0','5','6'))])    
        if self.reporte == '3':#Estadísticas Mensuales
            cand = self.env['custom.vlines'].search([('eta_date','>=',self.sdate),('eta_date','<=',self.edate),('etapa','in',('5','6'))])   
        if self.reporte == '4':#General
            cand = self.env['custom.vlines'].search([('eta_date','>=',self.sdate),('eta_date','<=',self.edate),('etapa','in',('0','5','6'))]) 
        if self.reporte == '5':#Incumplimiento
            cand = self.env['custom.vlines'].search([('eta_date','>=',self.sdate),('eta_date','<=',self.edate),('etapa','in',('0','6'))])
        if len(cand) == 0:
            raise ValidationError ('No hay contenedores registrados con fecha estimada en el rango provisto')
        if not self.reporte:
            raise ValidationError('Debe seleccionar un Reporte de la Lista') 
        if self.reporte == '1':
            wb =self.immex_rep(cand,self.sdate,self.edate,False)
            filename = 'Reporte IMMMEX.xlsx'
        if self.reporte == '2':
            wb =self.presidencia(cand,self.sdate,self.edate,False)
            filename = 'Reporte Presidencia.xlsx'
        if self.reporte == '3':
            wb =self.estadistica(cand,self.sdate,self.edate,False)
            filename = 'Reporte Estadística.xlsx'
        if self.reporte == '4':
            wb =self.general(cand,self.sdate,self.edate,False)
            filename = 'Reporte General.xlsx'
        if self.reporte == '5':
            wb =self.incumplimiento(cand,self.sdate,self.edate,False)
            filename = 'Reporte de Incumplimiento.xlsx'
        #Bajamos Excel
        with NamedTemporaryFile() as tmp: #graba archivo temporal
            wb.save(tmp.name) #graba el contenido del excel en tmp.name
            output = tmp.read()
        xlsx = {                            #características del archivo
                'name': filename,
                'type': 'binary',
                'res_model': 'selmrp.tmpexploit',
                'datas': base64.b64encode(output),  #aqui metemos el archivo generado y grabado
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                }
        inserted_id=self.env['ir.attachment'].create(xlsx) # creamos el link de download
        url='/web/content/%s?download=1' %(inserted_id.id)
        return {'type': 'ir.actions.act_url','name': filename,'url': url} #regresamos el link con el archivo         

    #-----------------------------------Facturación IMMEX------------------------------
    def immex_rep(self,cand,sdate,edate,immex):
        wb = Workbook() #creamos objeto
        ws = wb.active # inicializamos
        reng = 6 #indicador de renglon
        ws.title = "Facturación MC Immex" #titulo
        oper_data = []
        immex_data = []
        #traemos la imagen configurada en la compañia
        company_id = self.env.company
        if company_id.logo:
            buf_image= BytesIO(base64.b64decode(company_id.logo))
            img = Image(buf_image)
            img.anchor='A1'
            ws.add_image(img)
        ws.cell(4, 4).value = "Facturación MC Immex"
        ws.cell(4, 4).font = Font(size = "15")
        #--------------------Cabecera------------------------------------
        ws['A5'] = 'IMMEX'
        ws['B5'] = 'CATEGORIA'
        ws['C5'] = '#CONTENEDOR'
        ws['D5'] = 'OPERADORA'
        ws['E5'] = 'FECHA DE ETA'
        ws['F5'] = 'FECHA DE DESPACHO'
        for col in range (1,7):
            ws.cell(row=5, column=col).font = Font(color="FFFFFF")
            ws.cell(row=5, column=col).fill = PatternFill('solid', fgColor = '063970')
        for line in cand:
            #si no está en la lista de IMMEX no me importa para el Reporte
            if immex:
                if line.v_id.partner_id not in immex:
                    continue 
            #buscamos la línea origen 
            l_or = self.env['custom.task.line'].search([('task_id','=',line.v_id.id),('container_number','=',line.container_number)],limit = 1)
            _logger.error('task'+str(line.v_id.name)+'->Immex----->'+str(line.v_id.partner_id.name))
            if line.v_id.partner_id:
                ws.cell(row=reng, column=1).value = line.v_id.partner_id.name #IMMEX
            cat = ''
            if line.custom_category == '24':
                cat= 'IMMEX 24 hrs'
            if line.custom_category == '36':
                cat= 'IMMEX 36 hrs'
            ws.cell(row=reng, column=2).value = cat #categoría
            if line.container_number:
                ws.cell(row=reng, column=3).value = line.container_number #Contenedor 
            if l_or.operadora.name:                
                ws.cell(row=reng, column=4).value = str(l_or.operadora.name) #Operadora
            if line.eta_date:    
                ws.cell(row=reng, column=5).value = line.eta_date #Eta date
            if l_or.dispatch_date:    
                ws.cell(row=reng, column=6).value = str(l_or.dispatch_date)#Fecha de despacho
            #checamos si ya esta en oper_data
            esta = False
            for cd in oper_data:
                if cd[0] == str(l_or.operadora.name):
                    cd[1]= int(cd[1]) + 1
                    esta = True
                    break
            if not esta:
                oper_data.append([str(l_or.operadora.name),1])
            #checamos si ya esta en immex_data
            esta2 = False
            for cd in immex_data:
                if cd[0] == str(line.v_id.partner_id.name):
                    cd[1]= int(cd[1]) + 1
                    esta2 = True
                    break
            if not esta2:
                immex_data.append([str(line.v_id.partner_id.name),1])
            reng = reng + 1
        #--------------Contenedor por Operadora---------------------
        #raise ValidationError(str(oper_data))
        ws_chart = wb.create_sheet('Operadora')
        reng = 1
        ws_chart.cell(row=reng, column=1).value = 'Operadora'
        ws_chart.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in oper_data:
            reng = reng + 1
            ws_chart.cell(row=reng, column=1).value = row[0]
            ws_chart.cell(row=reng, column=2).value = row[1]
        pie = BarChart()
        mr = len(oper_data)+1
        labels = Reference(ws_chart, min_col=1, min_row=2, max_row=mr)
        oper_data = Reference(ws_chart, min_col=2, min_row=1, max_row=mr)
        pie.add_data(oper_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Contenedores por Operadora"
        ws_chart.add_chart(pie, "D2")            
        #-------------Fin de Char por Operadora-------------------------
        #--------------Contenedor por Immex---------------------
        #raise ValidationError(str(immex_data))
        ws_chart2 = wb.create_sheet('IMMEX')
        reng = 1
        ws_chart2.cell(row=reng, column=1).value = 'IMMEX'
        ws_chart2.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart2.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart2.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in immex_data:
            reng = reng + 1
            ws_chart2.cell(row=reng, column=1).value = row[0]
            ws_chart2.cell(row=reng, column=2).value = row[1]
        pie2 = BarChart()
        mr = len(immex_data)+1
        labels =    Reference(ws_chart2, min_col=1, min_row=2, max_row=mr)
        immex_data = Reference(ws_chart2, min_col=2, min_row=1,  max_row=mr)
        pie2.add_data(immex_data, titles_from_data=True)
        pie2.set_categories(labels)
        pie2.title = "Contenedores por IMMEX"
        ws_chart2.add_chart(pie2, "D2")            
        #-------------Fin de Char por immex-------------------------

        return wb

    #-----------------------------------Estadísticas Mensuales------------------------------
    def estadistica(self,cand,sdate,edate,immex):
        wb = Workbook() #creamos objeto
        ws = wb.active # inicializamos
        reng = 6 #indicador de renglon
        ws.title = "Estadísticas Mensuales" #titulo
        oper_data = []
        navi_data = []
        cate_data = []
        #traemos la imagen configurada en la compañia
        company_id = self.env.company
        if company_id.logo:
            buf_image= BytesIO(base64.b64decode(company_id.logo))
            img = Image(buf_image)
            img.anchor='A1'
            ws.add_image(img)
        ws.cell(4, 4).value = "Estadísticas Mensuales"
        ws.cell(4, 4).font = Font(size = "15")
        #--------------------Cabecera------------------------------------
        ws['A5'] = 'IMMEX'        
        ws['B5'] = 'CATEGORIA'
        ws['C5'] = '#CONTENEDOR'
        ws['D5'] = 'ID AGENTE ADUANAL'
        ws['E5'] = 'ID NAVIERA'
        ws['F5'] = 'ID FORWARDER'
        ws['G5'] = 'OPERADORA'
        ws['H5'] = 'FECHA DE ETA'
        ws['I5'] = 'FECHA DE DESPACHO'
        for col in range (1,10):
            ws.cell(row=5, column=col).font = Font(color="FFFFFF")
            ws.cell(row=5, column=col).fill = PatternFill('solid', fgColor = '063970')
        for line in cand:
            #si no está en la lista de IMMEX no me importa para el Reporte
            if immex:
                if line.v_id.partner_id not in immex:
                    continue 
            if line.v_id.partner_id:
                ws.cell(row=reng, column=1).value = str(line.v_id.partner_id.name) #IMMEX   
            #buscamos la línea origen 
            l_or = self.env['custom.task.line'].search([('task_id','=',line.v_id.id),('container_number','=',line.container_number)],limit = 1)
            if line.custom_category == '24':
                cat= 'IMMEX 24 hrs'
            if line.custom_category == '36':
                cat= 'IMMEX 36 hrs'
            ws.cell(row=reng, column=2).value = cat #categoría
            if line.container_number:
                ws.cell(row=reng, column=3).value = line.container_number #Contenedor 
            if l_or.agente_aduanal.name:
                ws.cell(row=reng, column=4).value = str(l_or.agente_aduanal.name) #Agente Aduanal
            if l_or.naviera:                
                ws.cell(row=reng, column=5).value = str(l_or.naviera) #Naviera 
            if l_or.forwarders.name:
                ws.cell(row=reng, column=6).value = str(l_or.forwarders.name) #Forwarder
            if l_or.operadora.name:
                ws.cell(row=reng, column=7).value = str(l_or.operadora.name) #Operadora
            if line.eta_date:
                ws.cell(row=reng, column=8).value = line.eta_date #Eta date
            if l_or.dispatch_date:
                ws.cell(row=reng, column=9).value = str(l_or.dispatch_date)#Fecha de despacho
            #------------->>>Datos para gráficas<<<<<-------------
            #-------------Contenedores por Operadora
            esta_oper = False
            for cd in oper_data:
                if cd[0] == str(l_or.operadora.name):
                    cd[1]= int(cd[1]) + 1
                    esta_oper = True
                    break
            if not esta_oper:
                oper_data.append([str(l_or.operadora.name),1])
            #------------------Contenedores por Naviera
            esta_navi = False
            for cd in navi_data:
                if cd[0] == str(l_or.naviera):
                    cd[1]= int(cd[1]) + 1
                    esta_navi = True
                    break
            if not esta_navi:
                navi_data.append([str(l_or.naviera),1])
            #-------------Contenedores por Categoría
            esta_cat = False
            for cd in cate_data:
                if cd[0] == str(cat):
                    cd[1]= int(cd[1]) + 1
                    esta_cat = True
                    break
            if not esta_cat:
                cate_data.append([cat,1])                        
            #------------->>>Datos para gráficas<<<<<-------------
            reng = reng + 1
        #------------------------>> gráficas<<----------------------
        #------------------------->>Contenedores por Operadora------
        ws_chart = wb.create_sheet('Operadora')
        reng = 1
        ws_chart.cell(row=reng, column=1).value = 'Operadora'
        ws_chart.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in oper_data:
            reng = reng + 1
            ws_chart.cell(row=reng, column=1).value = row[0]
            ws_chart.cell(row=reng, column=2).value = row[1]
        pie = BarChart()
        mr = len(oper_data)+1
        labels = Reference(ws_chart, min_col=1, min_row=2, max_row=mr)
        oper_data = Reference(ws_chart, min_col=2, min_row=1, max_row=mr)
        pie.add_data(oper_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Contenedores por Operadora"
        ws_chart.add_chart(pie, "D2")  
        #------------------------->>Contenedores por naviera------
        ws_chart2 = wb.create_sheet('Naviera')
        reng = 1
        ws_chart2.cell(row=reng, column=1).value = 'Naviera'
        ws_chart2.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart2.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart2.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in navi_data:
            reng = reng + 1
            ws_chart2.cell(row=reng, column=1).value = row[0]
            ws_chart2.cell(row=reng, column=2).value = row[1]
        pie = BarChart()
        mr = len(navi_data)+1
        labels = Reference(ws_chart2, min_col=1, min_row=2, max_row=mr)
        navi_data = Reference(ws_chart2, min_col=2, min_row=1, max_row=mr)
        pie.add_data(navi_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Contenedores por Naviera"
        ws_chart2.add_chart(pie, "D2")  
        #------------------------->>Contenedores por Categoría------
        ws_chart3 = wb.create_sheet('Categoría')
        reng = 1
        ws_chart3.cell(row=reng, column=1).value = 'Categoría'
        ws_chart3.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart3.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart3.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in cate_data:
            reng = reng + 1
            ws_chart3.cell(row=reng, column=1).value = row[0]
            ws_chart3.cell(row=reng, column=2).value = row[1]
        pie = BarChart()
        mr = len(cate_data)+1
        labels = Reference(ws_chart3, min_col=1, min_row=2, max_row=mr)
        cate_data = Reference(ws_chart3, min_col=2, min_row=1, max_row=mr)
        pie.add_data(cate_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Contenedores por Categoría"
        ws_chart3.add_chart(pie, "D2")  
        return wb

    #-----------------------------------Presidencia------------------------------
    def presidencia(self,cand,sdate,edate,immex):
        wb = Workbook() #creamos objeto
        ws = wb.active # inicializamos
        reng = 6 #indicador de renglon
        ws.title = "Presidencia" #titulo
        immex_data = []
        cate_data = []
        oper_data = []
        #traemos la imagen configurada en la compañia
        company_id = self.env.company
        if company_id.logo:
            buf_image= BytesIO(base64.b64decode(company_id.logo))
            img = Image(buf_image)
            img.anchor='A1'
            ws.add_image(img)
        ws.cell(4, 4).value = "Presidencia"
        ws.cell(4, 4).font = Font(size = "15")
        #--------------------Cabecera------------------------------------
        ws['A5'] = 'IMMEX'
        ws['B5'] = 'CATEGORIA'
        ws['C5'] = '#CONTENEDOR'
        ws['D5'] = 'TIPO CONTENEDOR'
        ws['E5'] = 'ID NAVIERA'
        ws['F5'] = 'OPERADORA'
        ws['G5'] = 'FECHA DE ETA'
        ws['H5'] = 'FECHA DE DESPACHO'
        for col in range (1,9):
            ws.cell(row=5, column=col).font = Font(color="FFFFFF")
            ws.cell(row=5, column=col).fill = PatternFill('solid', fgColor = '063970')
        for line in cand:
            #si no está en la lista de IMMEX no me importa para el Reporte
            if immex:
                if line.v_id.partner_id not in immex:
                    continue 
            #immex name
            if line.v_id.partner_id:
                ws.cell(row=reng, column=1).value = str(line.v_id.partner_id.name) #IMMEX   
            #buscamos la línea origen 
            l_or = self.env['custom.task.line'].search([('task_id','=',line.v_id.id),('container_number','=',line.container_number)],limit = 1)
            if line.custom_category == '24':
                cat= 'IMMEX 24 hrs'
            if line.custom_category == '36':
                cat= 'IMMEX 36 hrs'
            ws.cell(row=reng, column=2).value = cat #categoría
            if line.container_number:
                ws.cell(row=reng, column=3).value = line.container_number #Contenedor 
            if l_or.container_type_id.code:
                ws.cell(row=reng, column=4).value = l_or.container_type_id.code #tipo de Contenedor 
            if l_or.naviera:
                ws.cell(row=reng, column=5).value = str(l_or.naviera) #Naviera 
            if l_or.operadora.name:
                ws.cell(row=reng, column=6).value = str(l_or.operadora.name) #Operadora
            if line.eta_date:
                ws.cell(row=reng, column=7).value = line.eta_date #Eta date
            if l_or.dispatch_date:
                ws.cell(row=reng, column=8).value = str(l_or.dispatch_date)#Fecha de despacho
            #-------------Contenedores por Operadora
            esta_oper = False
            for cd in oper_data:
                if cd[0] == str(l_or.operadora.name):
                    cd[1]= int(cd[1]) + 1
                    esta_oper = True
                    break
            if not esta_oper:
                oper_data.append([str(l_or.operadora.name),1])
            #-------------Contenedores por Categoría
            esta_cat = False
            for cd in cate_data:
                if cd[0] == str(cat):
                    cd[1]= int(cd[1]) + 1
                    esta_cat = True
                    break
            if not esta_cat:
                cate_data.append([cat,1])
            #checamos si ya esta en immex_data
            esta_immex = False
            for cd in immex_data:
                if cd[0] == str(line.v_id.partner_id.name):
                    cd[1]= int(cd[1]) + 1
                    esta_immex = True
                    break
            if not esta_immex:
                immex_data.append([str(line.v_id.partner_id.name),1])
            reng = reng + 1                        
        #------------------------>> gráficas<<----------------------
        #------------------------->>Contenedores por Operadora------
        ws_chart = wb.create_sheet('Operadora')
        reng = 1
        ws_chart.cell(row=reng, column=1).value = 'Operadora'
        ws_chart.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in oper_data:
            reng = reng + 1
            ws_chart.cell(row=reng, column=1).value = row[0]
            ws_chart.cell(row=reng, column=2).value = row[1]
        pie = BarChart()
        mr = len(oper_data)+1
        labels = Reference(ws_chart, min_col=1, min_row=2, max_row=mr)
        oper_data = Reference(ws_chart, min_col=2, min_row=1, max_row=mr)
        pie.add_data(oper_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Contenedores por Operadora"
        ws_chart.add_chart(pie, "D2")  
        #------------------------->>Contenedores por IMMEX------
        ws_chart2 = wb.create_sheet('IMMEX')
        reng = 1
        ws_chart2.cell(row=reng, column=1).value = 'IMMEX'
        ws_chart2.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart2.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart2.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in immex_data:
            reng = reng + 1
            ws_chart2.cell(row=reng, column=1).value = row[0]
            ws_chart2.cell(row=reng, column=2).value = row[1]
        pie = BarChart()
        mr = len(immex_data)+1
        labels = Reference(ws_chart2, min_col=1, min_row=2, max_row=mr)
        immex_data = Reference(ws_chart2, min_col=2, min_row=1, max_row=mr)
        pie.add_data(immex_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Contenedores por IMMEX"
        ws_chart2.add_chart(pie, "D2")  
        #------------------------->>Contenedores por Categoría------
        ws_chart3 = wb.create_sheet('Categoría')
        reng = 1
        ws_chart3.cell(row=reng, column=1).value = 'Categoría'
        ws_chart3.cell(row=reng, column=2).value = 'Contenedores'
        for col in range (1,3):
            ws_chart3.cell(row=1, column=col).font = Font(color="FFFFFF")
            ws_chart3.cell(row=1, column=col).fill = PatternFill('solid', fgColor = '063970')
        for row in cate_data:
            reng = reng + 1
            ws_chart3.cell(row=reng, column=1).value = row[0]
            ws_chart3.cell(row=reng, column=2).value = row[1]
        pie = BarChart()
        mr = len(cate_data)+1
        labels = Reference(ws_chart3, min_col=1, min_row=2, max_row=mr)
        cate_data = Reference(ws_chart3, min_col=2, min_row=1, max_row=mr)
        pie.add_data(cate_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Contenedores por Categoría"
        ws_chart3.add_chart(pie, "D2")  

        return wb

   #-----------------------------------General------------------------------
    def general(self,cand,sdate,edate,immex):
        wb = Workbook() #creamos objeto
        ws = wb.active # inicializamos
        reng = 6 #indicador de renglon
        ws.title = "General" #titulo
        oper_data = []
        #traemos la imagen configurada en la compañia
        company_id = self.env.company
        if company_id.logo:
            buf_image= BytesIO(base64.b64decode(company_id.logo))
            img = Image(buf_image)
            img.anchor='A1'
            ws.add_image(img)
        ws.cell(4, 4).value = "General"
        ws.cell(4, 4).font = Font(size = "15")
        #--------------------Cabecera------------------------------------
        ws['A5'] = 'IMMEX'
        ws['B5'] = 'CATEGORIA'
        ws['C5'] = 'BL'
        ws['D5'] = '#CONTENEDOR'
        ws['E5'] = 'TIPO CONTENEDOR'
        ws['F5'] = 'ID AGENTE ADUANAL'
        ws['G5'] = 'ID NAVIERA'
        ws['H5'] = 'ID FORWARDERS'        
        ws['I5'] = 'OPERADORA'
        ws['J5'] = 'BUQUE'        
        ws['K5'] = 'NO. VIAJE'        
        ws['L5'] = 'FECHA DE ETA'
        ws['M5'] = 'FECHA PREVIO'
        ws['N5'] = 'FECHA DE DESPACHO'
        ws['O5'] = 'PREVIO'
        ws['P5'] = 'PESO'
        ws['Q5'] = 'PZA'
        ws['R5'] = 'EMBALAJE'

        for col in range (1,19):
            ws.cell(row=5, column=col).font = Font(color="FFFFFF")
            ws.cell(row=5, column=col).fill = PatternFill('solid', fgColor = '063970')
        for line in cand:
            #si no está en la lista de IMMEX no me importa para el Reporte
            if immex:
                if line.v_id.partner_id not in immex:
                    continue 
            #immex name
            if line.v_id.partner_id:
                ws.cell(row=reng, column=1).value = str(line.v_id.partner_id.name) #IMMEX   
            #buscamos la línea origen 
            l_or = self.env['custom.task.line'].search([('task_id','=',line.v_id.id),('container_number','=',line.container_number)],limit = 1)
            _logger.error('-------------------Reporte immex '+str(line.v_id.name)+' ---------- linea encontrada'+str(l_or))
            if line.custom_category == '24':
                cat= 'IMMEX 24 hrs'
            if line.custom_category == '36':
                cat= 'IMMEX 36 hrs'
            ws.cell(row=reng, column=2).value = cat #categoría
            if l_or.bl:
                ws.cell(row=reng, column=3).value = l_or.bl #BL
            if line.container_number:
                ws.cell(row=reng, column=4).value = line.container_number #Contenedor 
            if l_or.container_type_id.code:
                ws.cell(row=reng, column=5).value = l_or.container_type_id.code #tipo de Contenedor 
            if l_or.agente_aduanal:
                ws.cell(row=reng, column=6).value = str(l_or.agente_aduanal.name) #aGENTE ADUANAL
            if l_or.naviera:
                ws.cell(row=reng, column=7).value = str(l_or.naviera) #Naviera 
            if l_or.forwarders.name:
                ws.cell(row=reng, column=8).value = str(l_or.forwarders.name) #fORWARDERS
            if l_or.operadora.name:
                ws.cell(row=reng, column=9).value = str(l_or.operadora.name) #Operadora
            if l_or.buque:
                ws.cell(row=reng, column=10).value = l_or.buque #BUQUE
            if l_or.numero_viaje:
                ws.cell(row=reng, column=11).value = l_or.numero_viaje #NO.VIAJE
            if line.eta_date:
                ws.cell(row=reng, column=12).value = str(line.eta_date) #Eta date
            if l_or.previo_date:
                ws.cell(row=reng, column=13).value = str(l_or.previo_date) #FECHA PREVIO
            if l_or.dispatch_date:
                ws.cell(row=reng, column=14).value = str(l_or.dispatch_date)#Fecha de despacho
            #ws.cell(row=reng, column=14).value = line.eta_date #PREVIO
            if l_or.peso:
                ws.cell(row=reng, column=16).value = str(l_or.peso) #PESO
            if l_or.pieza:
                ws.cell(row=reng, column=17).value = l_or.pieza #PZA
            if l_or.packing_type_id.name:
                ws.cell(row=reng, column=18).value = l_or.packing_type_id.name #EMBALAJE
            reng = reng + 1
        return wb

   #-----------------------------------Incumplimiento------------------------------
    def incumplimiento(self,cand,sdate,edate,immex):
        wb = Workbook() #creamos objeto
        ws = wb.active # inicializamos
        reng = 6 #indicador de renglon
        ws.title = "Incumplimiento" #titulo
        oper_data = []
        #traemos la imagen configurada en la compañia
        company_id = self.env.company
        if company_id.logo:
            buf_image= BytesIO(base64.b64decode(company_id.logo))
            img = Image(buf_image)
            img.anchor='A1'
            ws.add_image(img)
        ws.cell(4, 4).value = "Reporte de Incumplimiento"
        ws.cell(4, 4).font = Font(size = "15")
        #--------------------Cabecera------------------------------------
        ws['A5'] = 'CONTENEDOR'
        #-------------Forwarder---------------------------------
        ws['B5'] = 'FORWARDER'
        ws['C5'] = 'CUMPLIMIENTO DOCUMENTACIÓN'
        ws['D5'] = 'CUMPLIMIENTO COSTO FLETE MARÍTIMO'
        #IMMEX
        ws['E5'] = 'IMMEX'
        ws['F5'] = 'PREALERTA REPORTE'
        #-------------aGENCIA aDUANAL
        ws['G5'] = 'AGENTE'
        ws['H5'] = 'REVALIDACIÓN DEL BL'        
        ws['I5'] = 'LIBERACIÓN DEL FOLIO'
        ws['J5'] = 'PROGRAMACIÓN DEL PREVIO'        
        ws['K5'] = 'PROGRAMACIÓN MANIOBRA DE CARGA'        
        #-------------TRANSPORTISTA
        ws['L5'] = 'TRANSPORTISTA'
        ws['M5'] = 'ASIGNACION PLACAS VEHICULO'
        ws['N5'] = 'ASIGNACION NOMBRE CONDUCTOR'
        #-------------eTA dATE
        ws['O5'] = 'FECHA ETA'

        for col in range (1,16):
            ws.cell(row=5, column=col).font = Font(color="FFFFFF")
            ws.cell(row=5, column=col).fill = PatternFill('solid', fgColor = '063970')
        for line in cand:
            #si no está en la lista de IMMEX no me importa para el Reporte
            if immex:
                if line.v_id.partner_id not in immex:
                    continue 
            if line.container_number:#CONTENEDOR
                ws.cell(row=reng, column=1).value = str(line.container_number) #CONTENEDOR
            if line.u_forwarder:#-------------------FORWARDER----------------------------------------
                ws.cell(row=reng, column=2).value = str(line.u_forwarder.name) #FORWARDER
            if line.vfdate:#CUMPLIMIENTO DOCUMENTACIÓN
                ws.cell(row=reng, column=3).value = str(line.vfdate) #CUMPLIMIENTO DOCUMENTACIÓN
                ws.cell(row=reng, column=4).value = str(line.vfdate) #CUMPLIMIENTO COSTO FLETE MARÍTIMO
            if line.v_id.partner_id:#IMMEX
                ws.cell(row=reng, column=5).value = str(line.v_id.partner_id.name) #immex name
            if line.u_aduanal:#-------------------Agente Aduanal----------------------------------------
                ws.cell(row=reng, column=7).value = str(line.u_aduanal.name) #immex name
            if line.vadate:
                ws.cell(row=reng, column=8).value = str(line.vdt_revalida) #revalidacion BL
                ws.cell(row=reng, column=9).value = str(line.vdt_folio) #liberacion folio
                ws.cell(row=reng, column=10).value = str(line.vdt_previo) #programacion del previo
            if line.vmdate:
                ws.cell(row=reng, column=11).value = str(line.vdt_maniobra) #PROGRAMACIÓN MANIOBRA DE CARGA
            if line.u_transportista:#-------------------Transportista----------------------------------------
                ws.cell(row=reng, column=12).value = str(line.u_transportista.name) #immex name
            if line.vtdate:
                ws.cell(row=reng, column=13).value = str(line.vtdate) #ASIGNACION PLACAS VEHICULO
                ws.cell(row=reng, column=14).value = str(line.vtdate) #ASIGNACION NOMBRE CONDUCTOR
            if line.eta_date:    
                ws.cell(row=reng, column=15).value = str(line.eta_date) #fecha eta
            reng = reng + 1
        return wb